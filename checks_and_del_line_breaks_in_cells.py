#!python3
import os
import sys
import logging
import re
import csv
import datetime
import time

import openpyxl

# logging.disable(level=logging.CRITICAL)
logdir = "/Users/mobile01/logs/"
log_name = "{}-checks_and_del.log".format(datetime.date.today())
logging.basicConfig(
    level=logging.INFO,
    format=' %(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("{0}/{1}".format(logdir, log_name)),
        logging.StreamHandler()
    ])
logger = logging.getLogger()

logger.debug('Start of program')
re_file = re.compile(r'(\d+-\d+-)?(\w+)(-\S+)?.\S+')
data_dir = '/Users/mobile01/%s' % datetime.date.today()

if not os.path.exists(data_dir):
    os.mkdir(data_dir)

Cols_num = {
    'p_customer': '41',
    'p_customer_qianke': '12',
    'p_customer_car': '21',
    'p_complain': '23',
    'o_car': '12',
    'o_car_order': '35',
    'i_parts': '20',
    'i_parts_inventory': '6'
}

Cell_not_null = {
    'p_customer': [1, 2, 3, 33, 35, 36, 37],
    'p_customer_qianke': [1, 2, 3, 4, 5, 8, 12],
    'p_customer_car': [1, 2, 5, 6, 13, 14, 15],
    'p_complain': [1, 2, 3, 4, 6, 7, 8, 10, 12, 13, 14, 16, 23],
    'o_car': [1, 2, 3, 4, 6, 7, 8, 9, 10, 12],
    'o_car_order': [1, 2, 3, 5, 6, 8],
    'i_parts': [1, 2, 3, 5, 6, 7, 9, 10, 11, 12, 17, 18],
    'i_parts_inventory': [1, 2, 3, 4, 5, 6]
}

Cell_not_date = {
    'p_customer': [8, 36],
    'p_customer_qianke': [3],
    'p_customer_car': [9, 10, 12, 14, 17, 20],
    'p_complain': [1, 6, 17, 19],
    'o_car': [3, 8, 9],
    'o_car_order': [2],
    'i_parts': [17, 19],
}
# 传入要处理的Excel文件路径：
if len(sys.argv) > 1:
    excel_dir = sys.argv[1]
else:
    print('ERROR:请传入要检查的Excel所在目录！')
    exit(2)


def checks_max_col_num(excel_filename, re_file, col_num):
    no = int(Cols_num.get(re_file))
    if col_num != no:
        print('ERROR:%s 文件列数错误！请检查相关文件！现列数为: %s' % (excel_filename, col_num))
        exit(3)


def checks_not_null(excel_filename, re_file_, cols_num, data_, line_num_):
    if cols_num in Cell_not_null.get(re_file_):
        if data_ == 0:
            return 0
        elif not data_:
            print('ERROR:%s 文件 %s行%s列 不能为空！' % (excel_filename, line_num_, cols_num))
            exit(4)

    if data_ and line_num_ != 1 and not isinstance(data_, datetime.datetime):
        if re_file in Cell_not_date.keys():
            if cols_num in Cell_not_date.get(re_file_):
                if not is_valid_date(data_):
                    print('ERROR:%s 文件 %s行%s列的值非法，不是日期格式！内容:"%s"!'
                          % (excel_filename, line_num_, cols_num, data_))


def is_valid_date(self):
    """判断是否是一个有效的日期字符串"""
    try:
        if ":" in self:
            if "." in self:
                time.strptime(self, "%Y-%m-%d %H:%M:%S.%f")
            else:
                time.strptime(self, "%Y-%m-%d %H:%M:%S")
        elif "/" in self:
            time.strptime(self, "%Y/%m/%d")
        else:
            time.strptime(self, "%Y-%m-%d")
        return True
    except ValueError:
        return False


def del_line_breaks_colons(self):
    if self is None:
        self = ''
        return self
    else:
        if type(self) == float:
            if int(self) == self:
                self = int(self)
        self = str(self)
        self = self.replace('\n', ' ')
        self = self.replace(',', '，')
        self = self.replace("'", " ")
        self = self.replace("\"", " ")
        return self


if os.path.exists(excel_dir):
    for folderName, subfolders, filenames in os.walk(excel_dir):
        files = [f for f in filenames if not f[0] == '.' and not f[0] == '~' and f.endswith('.xlsx')]
        subfolders[:] = [d for d in subfolders if not d[0] == '.']
        os.chdir(folderName)
        for filename in files:
            logger.debug(filename)
            table_object = re_file.search(filename)
            re_table = table_object.group(2)
            logger.debug(re_table)
            if re_table not in Cols_num.keys():
                print('%s 文件名不正确或者不符合规则！请检查修改！' % filename)
                exit(2)
            data = openpyxl.load_workbook(filename=filename, read_only=True)
            wb = openpyxl.Workbook()
            w_data = wb.active
            r_data = data.active
            cols = r_data.max_column
            checks_max_col_num(filename, re_table, cols)
            line_num = 0
            l_row = []
            for row in r_data.iter_rows():
                line_num += 1
                l_row = [del_line_breaks_colons(f.value) for f in row]
                for num in range(cols):
                    cell = l_row[num]
                    checks_not_null(filename, re_table, num + 1, cell, line_num)
                w_data.append(l_row)
            wb.save(filename)

if os.path.exists(excel_dir):
    for folderName, subfolders, filenames in os.walk(excel_dir):
        files = [f for f in filenames if not f[0] == '.' and f.endswith('.xlsx')]
        subfolders[:] = [d for d in subfolders if not d[0] == '.']
        os.chdir(folderName)
        for filename in files:
            logger.debug(filename)
            temp_table = filename.replace('xlsx', 'csv')
            logger.debug(temp_table)
            data = openpyxl.load_workbook(filename)
            table = data.active
            csv_file = open('%s/%s' % (data_dir, temp_table), 'w', newline="", encoding='utf-8-sig')
            c = csv.writer(csv_file)
            for r in table.rows:
                c.writerow([cell.value for cell in r])
            csv_file.close()

print('清洗完毕')
logger.debug('End of program')
